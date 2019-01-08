import numpy as np  # 有时用来判断列表维度，若无该库，请把 assert np.array(list).ndim == num 注释！


# ==============================================================================
# 几个调用实例：
# result = read_xls_file('./info.xlsx', name_list=['姓名', '班级'])  # 返回字典型、表单切片的数据
# result = read_xls_file('./info.xls', sheet_var='StudentInfo', name_list=['姓名', '班级'],
#          category='col', begin=0, export_type='list')  # 返回列表型、StudentInfo表单、指定列的数据（还包括了列名）
# 
# write_xls_file([[[1,2,3], [1,2]], [[4]]], './test.xlsx', sheet_list=['tmp1', 'tmp2'])
# write_xls_file([[[1,2,3], [1,2], [4]]], './test.xls', category='col')
#
# result = read_csv_file('./info.csv', name_list=['姓名', '班级'])  # 返回字典型、表单切片的数据
#
# write_csv_file([['1','2','3'], [1,2], [4]], './test.csv', mode='w')
# write_csv_file([['1','2','3'], [1,2], [4]], './test.csv', mode='a', category='col')
#
# result = read_txt_file('./test.txt', code_type='utf-16')
# result = read_txt_file('./test.txt', n_num=10)  # 读取以'\n'为分割方式的前10行文本
#
# write_txt_file(['I love nlp', 'hello world'], './test.txt')
# ==============================================================================
def read_xls_file(file_path, sheet_var=0, name_list=[], category='', begin=1, export_type='dict'):
    """
    read only one .xls or .xlsx file and obtain the specific content

    :param file_path: string, the path of this file
    :param sheet_var: string or int, the name or index of current sheet
    :param name_list: 1d list, consists of names
    :param category: string, process by rows or columns or table slice(default and recommend)
    :param begin: int, the beginning position of a row or column list, but invalid when name_list is []
    :param export_type: string, denote the type of data
    :return: dict or 2d list, result from this excel
    """

    assert np.array(name_list).ndim == 1
    assert isinstance(category, str)
    assert isinstance(begin, int)
    assert isinstance(export_type, str)

    result = {} if export_type == 'dict' else []
    try:
        import xlrd
        er = xlrd.open_workbook(file_path)
        if isinstance(sheet_var, int):
            s = er.sheet_by_index(sheet_var)
        else:
            s = er.sheet_by_name(sheet_var)

        if name_list:  # 有指定名称，说明要导出 n 行或 n 列
            if category == 'col':
                for name in name_list:
                    pos = -1
                    for j in range(s.ncols):
                        if str(s.cell_value(0, j)).strip() == name:
                            pos = j
                            break

                    if pos >= 0:
                        if isinstance(result, dict):
                            result[name] = tuple(s.col_values(pos)[begin:])
                        else:  # list
                            result.append(tuple(s.col_values(pos)[begin:]))

            elif category == 'row':
                for name in name_list:
                    pos = -1
                    for i in range(s.nrows):
                        if str(s.cell_value(i, 0)).strip() == name:
                            pos = i
                            break

                    if pos >= 0:
                        if isinstance(result, dict):
                            result[name] = tuple(s.row_values(pos)[begin:])
                        else:  # list
                            result.append(tuple(s.row_values(pos)[begin:]))

            else:  # table slice [(col1, col2, col3),] 每一行元素按照 name_list 的顺序排列,每一行是元组
                tmp_data = read_xls_file(file_path=file_path, sheet_var=sheet_var, name_list=name_list,
                                         category='col', begin=begin, export_type='list')
                tmp_list = list(zip(*tmp_data))

                if isinstance(result, dict):
                    result['slice'] = tmp_list
                else:
                    result = tmp_list

        else:  # 没有指定名称，返回全部表格数据
            if category == 'col':  # 返回多列构成的二维列表
                tmp_list = [tuple(s.col_values(j)) for j in range(s.ncols)]
            else:    # category != 'col', 返回多行构成的二维列表
                tmp_list = [tuple(s.row_values(i)) for i in range(s.nrows)]

            if isinstance(result, dict):
                sheet_name = er.sheet_names()[0]
                result[sheet_name] = tmp_list
            else:
                result = tmp_list

    except Exception as err:
        print(err)

    return result


def write_xls_file(info_list, file_path, sheet_list=['Sheet1'],  category='row'):
    """
    write .xls or .xlsx file with organized input

    :param info_list: 3d list, consists of all information (sheet, table, content)
    :param file_path: string, the path of this file
    :param sheet_list: 1d list, consists of names of these sheets
    :param category: string, insert by rows or columns
    :return: no essential return value
    """

    def transpose(sub_info_list):
        """
        from [row, col] to [col, row], both input and output are 2d list
        """
        col_num = len(sub_info_list)
        size = max([len(tmp) for tmp in sub_info_list])

        for j in range(col_num):  # padding
            length = len(sub_info_list[j])
            if length < size:
                sub_info_list[j].extend(['' for _ in range(size-length)])

        rev_list = [[sub_info_list[j][i] for j in range(col_num)] for i in range(size)]

        return rev_list

    assert isinstance(info_list, list)
    assert np.array(sheet_list).ndim == 1
    assert isinstance(category, str)
    assert len(info_list) == len(sheet_list)

    try:
        if file_path[-5:].find('.xlsx') != -1:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            font = Font(u'宋体', size=11)
            wb = Workbook()

            for k in range(len(sheet_list)):
                if k == 0:
                    sheet = wb.active
                else:
                    sheet = wb.create_sheet(sheet_list[k])

                sheet.title = sheet_list[k]
                sheet.font = font

                if category == 'row':
                    tmp_list = info_list[k]
                else:  # col
                    tmp_list = transpose(info_list[k])

                for tmp in tmp_list:
                    sheet.append(tmp)

            if sheet_list:
                wb.save(file_path)
                print('write excel successfully')
            else:
                print('warning: no sheet')

        elif file_path[-5:].find('.xls') != -1:
            import xlwt
            style = xlwt.XFStyle()  # 格式信息
            font = xlwt.Font()  # 字体基本设置
            font.name = u'宋体'
            font.color = 'black'
            font.height = 220  # 字体大小，220就是11号字体
            style.font = font

            alignment = xlwt.Alignment()  # 设置字体在单元格的位置
            alignment.horz = xlwt.Alignment.HORZ_CENTER  # 水平方向
            alignment.vert = xlwt.Alignment.VERT_CENTER  # 竖直方向
            style.alignment = alignment

            wb = xlwt.Workbook()

            for k in range(len(sheet_list)):
                s = wb.add_sheet(sheet_list[k])

                if category == 'row':  # 一行一行写
                    for i in range(len(info_list[k])):
                        for j in range(len(info_list[k][i])):
                            s.write(i, j, info_list[k][i][j], style)

                else:  # col, 一列一列写
                    for j in range(len(info_list[k])):
                        for i in range(len(info_list[k][j])):
                            s.write(i, j, info_list[k][j][i], style)

            if sheet_list:
                wb.save(file_path)
                print('write excel successfully')
            else:
                print('warning: no sheet')

        else:
            print('warning: no such file type')

    except Exception as err:
        print(err)


def read_csv_file(file_path, name_list=[], category='', begin=1, export_type='dict'):
    """
    read only one .csv file and obtain the specific content

    :param file_path: string, the path of this file
    :param name_list: 1d list, consists of names
    :param category: string, process by rows or columns
    :param begin: int, the beginning position of a row or column list, but invalid when name_list is []
    :param export_type: string, denote the type of data
    :return: dict or 2d list, result from this csv file
    """

    assert np.array(name_list).ndim == 1
    assert isinstance(category, str)
    assert isinstance(begin, int)
    assert isinstance(export_type, str)

    result = {} if export_type == 'dict' else []
    try:
        import csv
        with open(file_path, 'r') as f:
            info_list = list(csv.reader(f))  # iter to list

        if not info_list:
            return result

        row_num = len(info_list)
        col_num = len(info_list[0])  # csv模块读取的每行元素具有相同的长度
        if name_list:  # 有指定名称，说明要导出 n 行或 n 列
            if category in ['col', 'row']:
                if category == 'col':
                    info_list = [tuple([info_list[i][j] for i in range(row_num)]) for j in range(col_num)]
                    length = col_num
                else:   # category == 'row'
                    info_list = [tuple(tmp) for tmp in info_list]
                    length = row_num

                for name in name_list:
                    pos = -1
                    for k in range(length):
                        if info_list[k][0].strip() == name:
                            pos = k
                            break

                    if pos >= 0:
                        if isinstance(result, dict):
                            result[name] = info_list[pos][begin:]
                        else:  # list
                            result.append(info_list[pos][begin:])

            else:  # table slice [(col1, col2, col3),] 每一行元素按照 name_list 的顺序排列,每一行是元组
                tmp_data = read_csv_file(file_path=file_path, name_list=name_list,
                                         category='col', begin=begin, export_type='list')
                tmp_list = list(zip(*tmp_data))

                if isinstance(result, dict):
                    result['slice'] = tmp_list
                else:
                    result = tmp_list

        else:  # 没有指定名称，返回全部表格数据
            if category == 'col':  # 返回多列构成的二维列表
                info_list = [tuple([info_list[i][j] for i in range(row_num)]) for j in range(col_num)]
            else:
                info_list = [tuple(tmp) for tmp in info_list]

            if isinstance(result, dict):
                result['csv_data'] = info_list
            else:
                result = info_list

    except Exception as err:
        print(err)

    return result


def write_csv_file(info_list, file_path, mode='w', category='row'):
    """
    write .csv file with organized input

    :param info_list: 2d list, consists of all information
    :param file_path: string, the path of this file
    :param mode: string, including w and a, generally we use w
    :param category: string, insert by rows or columns
    :return: no essential return value
    """

    def transpose(input_list):
        col_num = len(input_list)
        size = max([len(tmp) for tmp in input_list])
        for j in range(col_num):  # padding
            length = len(input_list[j])
            if length < size:
                input_list[j].extend(['' for _ in range(size - length)])

        rev_list = [[input_list[j][i] for j in range(col_num)] for i in range(size)]

        return rev_list

    assert isinstance(info_list, list)
    assert isinstance(mode, str)
    assert isinstance(category, str)

    try:
        import csv
        if category == 'row':
            tmp_list = info_list
        else:    # col
            tmp_list = transpose(info_list)

        with open(file_path, mode, newline='') as wr:
            writer = csv.writer(wr)
            writer.writerows(tmp_list)

    except Exception as err:
        print(err)


def read_txt_file(file_path, n_num=-1, code_type='utf-8'):
    """
    read .txt files, get all text or the previous n_num lines

    :param file_path: string, the path of this file
    :param n_num: int, denote the row number decided by \n, but -1 means all text
    :param code_type: string, the code of this file
    :return: string, text
    """

    with open(file_path, 'r', encoding=code_type) as f:
        if n_num <= 0:
            text = f.read().strip()
        else:  # n_num > 0
            text = '\n'.join([f.readline() for _ in range(n_num)])

    return text


def write_txt_file(text_list, file_path, mode='w', code_type='utf-8'):
    """
    write text to .txt files in the disk, attention: text_list has been organized well by 1d list

    :param text_list: 1d list, it consists of organized information
    :param file_path: string, the path of this file
    :param mode: string, including w and a, generally we use w
    :param code_type: string, the code of this file
    :return: no essential return value
    """

    assert np.array(text_list).ndim == 1
    assert isinstance(mode, str)

    with open(file_path, mode, encoding=code_type) as wr:
        wr.write('\n'.join(text_list) + '\n')
                            




